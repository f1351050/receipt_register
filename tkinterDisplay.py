import tkinter as tk
from PIL import Image, ImageTk, ImageOps
import tkinter.font as f
from tkinter import messagebox
from tkinter import filedialog
import pandas as pd
import tkinter.ttk as ttk
import numpy as np 
import ocr_receipt
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil
import numpy as np
import os
import re
import win32com.client

#!
# todo↓たまにできない
# todo同じ中分類・商品分類を2回以上登録しようとすると同じ番号になってしまう
# todo 削除
# todo　上↓がおかしい
# todo 最終category登録のウィンドウでも正式名所を登録できるようにする
#todo　category登録でのエラー　➡　検索窓で入力して、noも反映。
#todo　✔ボックスつけて反映文字
#todo 店舗登録
#?
#*
#>

#master=tk.Tk() self.master=root=tk.Tk()
class Application(tk.Frame):
    def __init__(self,master=None):
        super().__init__(master)
        self.excel_file_path=r"C:\Users\f1351\Desktop\rireki.xlsx"
        self.rireki_excel_get()#rireki,category,tempoExcel表をupdate
        self.pack()
        self.sub_win = None #category登録の際のサブウィンドウ初期
        self.sub_win_txt =None
        self.sub_win_noCheck=None
        self.increase=0 #画面上で増加した文の行の数（初期）
        # self.disabled_list=[]#display上で削除した行
        self.active_list=[]#deleteしてるしていないかの判定
        self.base_df_list=[] #元のrfdf表を保持
        
        self.master.title("レシート")
        self.master.geometry("950x800")

        # self.create_menu()
        #ステータスラベル
        frame_statusbar = tk.Frame(self.master, relief = tk.SUNKEN, bd = 2)
        self.statusbar = tk.Label(frame_statusbar, text = "StatusLabel")
        self.statusbar.pack(side = tk.LEFT)
        frame_statusbar.pack(side = tk.BOTTOM, fill = tk.X)
        

        #画像表示
        self.canvas=tk.Canvas(self.master,width=200,bg = "black")
        self.canvas.pack(side=tk.LEFT,fill = tk.Y)
        self.mouse_image()

        # 右枠下 
        self.frame_raight_bottom=tk.Frame(self.master,height=200,relief = tk.SUNKEN, bd = 3)
        self.frame_raight_bottomLeft=tk.Frame(self.frame_raight_bottom,height=200)
        self.frame_raight_bottomRight=tk.Frame(self.frame_raight_bottom,height=200,relief=tk.RIDGE,bd =2)
        self.category_kensaku()
        self.create_rireki_label()
        self.frame_raight_bottom.pack(side=tk.BOTTOM,fill = tk.X)
        self.frame_raight_bottomLeft.pack(side=tk.LEFT)
        self.frame_raight_bottomRight.pack(fill=tk.BOTH)
        #合計金額表示枠
        self.frame_sum=tk.Frame(self.master)
        self.kingaku_sum()
        self.frame_sum.pack(side=tk.BOTTOM,fill = tk.X)
        #ファイル名表示
        self.frame_right_best_top=tk.Frame(self.master,width=750)
        self.filename_dis()
        self.frame_right_best_top.pack(fill = tk.X)

        #日付、店舗名　表示部分
        self.frame_right_top=tk.Frame(self.master,width=750,height=100)
        self.frame_right_top.pack(fill = tk.X)

        #商品名等の抽出表示（scroll付きフレーム）
        self.raight_canvas = tk.Canvas(self.master,width=500)
        self.frame_right=tk.Frame(self.raight_canvas)
        self.frame_right.bind_all("<KeyPress>",self.key_push)

        self.frame_raight_bottomLeft.bind_all("<Control-Down>",self.kensaku_entry_focus)
        self.frame_raight_bottomLeft.bind_all("<Control-Left>",self.key_left)
        self.frame_raight_bottomLeft.bind_all("<Control-Right>",self.key_right)

        scroll = tk.Scrollbar(self.master, orient=tk.VERTICAL, command=self.raight_canvas.yview) 
        self.frame_right.bind("<Configure>",lambda e: self.raight_canvas.configure(scrollregion=self.raight_canvas.bbox("all")))
        self.raight_canvas.configure(yscrollcommand=scroll.set) 
        scroll.pack(side="right",fill="y")
        self.raight_canvas.bind("<MouseWheel>",self.mouse_y_scroll)
        self.category_label()
        self.rfdf=self.syokiRFDF()
        self.text_dis()

        self.frame_right.bind_all("<Button-1>",self.entry_click)
        self.frame_right.bind_all("<Double-1>",self.no_double)
        self.raight_canvas.pack(fill = tk.BOTH,expand = True)
        self.raight_canvas.create_window((0,0), window=self.frame_right, anchor="nw")
        self.create_menu()
    #*-------------------------------------------------------------------------------
    #*menubar作成
    #*-------------------------------------------------------------------------------
        #!ラベル_buttonの作成
    def create_menu(self):
        menubar=tk.Menu(self)
        self.master.config(menu = menubar)
        #file開くメニュー
        menu_file=tk.Menu(menubar,tearoff = False)
        menubar.add_cascade(label="ファイル",menu = menu_file)
        menu_file.add_command(label='画像表示',command=self.file_path_get,  accelerator="Ctrl+O")
        menu_file.bind_all("<Control-o>", self.file_path_get)
        
        menu_file.add_command(label='categoryを保存',command=self.advance_register,  accelerator="Ctrl+Q")
        menu_file.bind_all("<Control-q>", self.advance_register)

        menu_file.add_command(label='rirekiを保存',command=self.check_dis_register,  accelerator="Ctrl+S")
        menu_file.bind_all("<Control-s>",self.check_dis_register)
        menu_file.add_command(label='Excel表をUpdate',
                              command=lambda:[self.rireki_excel_get(),
                                            self.category_kensaku(),
                                            messagebox.showinfo("メッセージ","Excel表をアップデートしました")],
                              accelerator="Ctrl+E")
        menu_file.bind_all("<Control-e>",
                           lambda event:[self.rireki_excel_get(event),
                                        self.category_kensaku(event),
                                        messagebox.showinfo("メッセージ","Excel表をアップデートしました")])
        menu_file.add_command(label='Excelを開く',command=self.open_excel,  accelerator="F1")
        menu_file.bind_all("<Key-F1>", self.open_excel)

        #文字抽出 before_extraction_verMoziretu
        date_extract=tk.Menu(menubar,tearoff = False)
        menubar.add_cascade(label="文字抽出",menu = date_extract)
        date_extract.add_command(label='(整形前)文字列表示',command=self.before_extraction,  accelerator="Ctrl+M")
        date_extract.bind_all("<Control-m>",self.before_extraction)
        date_extract.add_command(label='(空白を除去)文字列表示',command=self.before_extraction_verMoziretu,  accelerator="Ctrl+K")
        date_extract.bind_all("<Control-k>",self.before_extraction_verMoziretu)
        date_extract.add_command(label='抽出',command=self.ocr,  accelerator="Ctrl+I")
        date_extract.bind_all("<Control-i>",self.ocr)

        #編集
        hensyuu=tk.Menu(menubar,tearoff = False)
        menubar.add_cascade(label="編集",menu = hensyuu)
        hensyuu.add_command(label='追加',command=self.tuika_index,  accelerator="Ctrl+A")
        hensyuu.bind_all("<Control-a>",self.tuika_index)
        hensyuu.add_command(label='元に戻す(商品名のみ)',command=self.result_initial_value,accelerator="Ctrl+Z")
        hensyuu.bind_all("<Control-z>",self.result_initial_value)
        hensyuu.add_command(label='削除',command=self.current_delete,  accelerator="Ctrl+D")
        hensyuu.bind_all("<Control-d>",self.current_delete)

        # そのほか
        hoka=tk.Menu(menubar,tearoff = False)
        menubar.add_cascade(label="その他",menu = hoka)
        hoka.add_command(label='category登録有無確認',command=self.update_unregistered,  accelerator="Ctrl+U")
        hoka.bind_all("<Control-u>",self.update_unregistered)
        hoka.add_command(label='値引き調整',command=self.nebiki(i=self.extractvar.get()),  accelerator="Ctrl+N")
        hoka.bind_all("<Control-n>",lambda event:[self.nebiki(i=self.extractvar.get()),
                                                  messagebox.showinfo("メッセージ","値引き前または後の金額に訂正しました")])
        hoka.add_command(label='合計金額の再調整',command=self.sum_process,  accelerator="Ctrl+G")
        hoka.bind_all("<Control-g>",lambda event:[self.sum_process(),
                                                  messagebox.showinfo("メッセージ","合計金額を訂正しました")])
        hoka.add_command(label='アプリを終了する',command=lambda:self.master.destroy(),
                                                          accelerator="F5")
        hoka.bind_all("<Key-F5>",lambda event:self.master.destroy())
    
    #*-------------------------------------------------------------------------------
    #* tkinterでの表記
    #* -------------------------------------------------------------------------------
    #!ファイルの表示
    def filename_dis(self):
        label_filename=tk.Label(self.frame_right_best_top,text='ファイルパス：')
        label_filename.pack(side =tk.LEFT)
        self.file_name_dis=tk.Entry(self.frame_right_best_top,width=80)
        self.file_name_dis.pack(side = tk.LEFT)
        file_change=tk.Button(self.frame_right_best_top,text="...",command=self.file_change_process)
        file_change.pack(side = tk.LEFT,pady=5)
    
    def file_change_process(self):
        provisional_filename=self.file_name_dis.get()
        if os.path.exists(provisional_filename)==False:
            messagebox.showerror('エラー','このファイルパスは存在しません')
            return
        self.filename=provisional_filename
        self.set_image()

    #!ラベル_buttonの作成
    def category_label(self):
        label_hizuke=tk.Label(self.frame_right_top,text='日付',anchor=tk.E)
        label_tempo=tk.Label(self.frame_right_top,text='店舗名',anchor=tk.E)
        label_daibunrui=tk.Label(self.frame_right,text='大分類',anchor=tk.S)
        label_tyuubunrui=tk.Label(self.frame_right,text='中分類',anchor=tk.S)
        label_no=tk.Label(self.frame_right,text='No',anchor=tk.S)
        label_syouhinmei=tk.Label(self.frame_right,text='商品名',anchor=tk.S)
        label_suuryou=tk.Label(self.frame_right,text='数量',anchor=tk.S)
        label_tani=tk.Label(self.frame_right,text='単位',anchor=tk.S)
        label_kingaku=tk.Label(self.frame_right,text='金額',anchor=tk.S)
        label_memo=tk.Label(self.frame_right,text='メモ',anchor=tk.S)

        label_hizuke.place(relx=0.12, rely=0.4)
        label_tempo.place(relx=0.1, rely=0.7)
        label_daibunrui.grid(row=3,column=1,padx=2,pady=2)
        label_tyuubunrui.grid(row=3,column=2,padx=2,pady=2)
        label_no.grid(row=3,column=3,padx=2,pady=2)
        label_syouhinmei.grid(row=3,column=4,padx=2,pady=2)
        label_suuryou.grid(row=3,column=5,padx=2,pady=2)
        label_tani.grid(row=3,column=6,padx=2,pady=2)
        label_kingaku.grid(row=3,column=7,padx=2,pady=2)
        label_memo.grid(row=3,column=8,padx=2,pady=2)

    #!日付と店舗名の表示,ファイルパスを表示
    def hizuke_tempo_display(self,list_tempo):
        df=self.rfdf
        self.tempo_combobox=ttk.Combobox(self.frame_right_top,values=list_tempo,width=30)
        self.hizuke=tk.Entry(self.frame_right_top,width=35)
        self.hizuke.insert(tk.END,df.iloc[0]['hizuke'])
        self.tempo_combobox.set(df.iloc[0]['tempo'])
        self.hizuke.place(relx=0.2, rely=0.4)
        self.tempo_combobox.place(relx=0.2, rely=0.7)

    #!Excelからcategoryの抽出
    def tempo_category_list(self):
        cate_df=self.category_df
        tempo_df=self.tempo_df
        list_dai_bunrui=cate_df['dai_bunrui'][~cate_df.duplicated(subset='dai_bunrui')].values.tolist()
        list_tyuu_bunrui=cate_df['tyuu_bunrui'][~cate_df.duplicated(subset='tyuu_bunrui')].values.tolist()
        list_tani=cate_df['tani'][~cate_df.duplicated(subset='tani')].values.tolist()
        list_tempo=tempo_df['tempo_mei'][~tempo_df.duplicated(subset='tempo_mei')].values.tolist()
        return list_tempo,list_dai_bunrui,list_tyuu_bunrui,list_tani
    
    #!OCR ocr_receiptで抽出した文字を表示
    def ocr(self,event=None):
        if not self.filename:
            return
        # 初期化させる
        if len(self.rfdf)+self.increase>=1:
            for i in range(len(self.rfdf)+self.increase):
                globals()[f'puru{i}'].destroy()
                globals()[f'daibunrui{i}'].destroy()
                globals()[f'tyubunrui{i}'].destroy()
                globals()[f'no{i}'].destroy()
                globals()[f'syokuhinmei{i}'].destroy()
                globals()[f'kazu{i}'].destroy()
                globals()[f'tani{i}'].destroy()
                globals()[f'unregistered{i}'].destroy()
                globals()[f'kingaku{i}'].destroy()
                globals()[f'memo{i}'].destroy()
            self.rfdf=self.syokiRFDF()
            self.active_list=[]
            self.base_df_list=[]
        self.statusbar["text"] = " OCR_NOW! "
        ocr=ocr_receipt.ocr(self.filename,self.df,self.category_df,self.tempo_df)
        self.increase=0
        try:
            self.txt,self.moziretu,self.rfdf=ocr.main()
        except Exception as e:
            messagebox.showerror('エラー', "抽出できませんでした。抽出前の文字列は表示可能("+str(e)+")")
            self.txt=ocr.ocr(self.filename)
            self.rfdf=self.syokiRFDF()
        if len(self.rfdf)==0:
            messagebox.showerror('エラー', "抽出できませんでした。抽出前の文字列は表示可能(不明なエラー)")
            self.rfdf=self.syokiRFDF()
        self.text_dis()
        self.purchase_history()
        self.sum_process()
    #!読み取り結果
    def text_dis(self):
        list_tempo,list_dai_bunrui,list_tyuu_bunrui,list_tani=self.tempo_category_list()
        self.hizuke_tempo_display(list_tempo)
        kazu_list=[1,2,3,4,5,6,7,8,9,10]
        # チェック有無変数
        self.extractvar = tk.IntVar()

        self.color='black'
        for i in range(len(self.rfdf)):
            self.insert_display(i,i,self.rfdf,list_dai_bunrui,list_tyuu_bunrui,kazu_list,list_tani)
        self.input_kensaku_entry()
        globals()[f'syokuhinmei{0}'].focus_set()
    
    #!syokirireki
    def create_rireki_label(self):
        self.r_label=tk.Label(self.frame_raight_bottomRight)
        self.r_label.grid(row=0,column=1,columnspan = 3,padx=2,pady=2)
        self.r_label['text']='購入履歴はありません'
        self.r_shmei=tk.Label(self.frame_raight_bottomRight)
        self.r_shmei.grid(row=1,column=1,columnspan = 3,padx=2,pady=2)

        for i in range(7):
            globals()[f'r_hizuke{i}']=tk.Label(self.frame_raight_bottomRight)
            globals()[f'r_value{i}']=tk.Label(self.frame_raight_bottomRight)
            globals()[f'r_tani{i}']=tk.Label(self.frame_raight_bottomRight)
            globals()[f'r_kingaku{i}']=tk.Label(self.frame_raight_bottomRight)
            
            globals()[f'r_hizuke{i}'].grid(row=i+2,column=1,padx=2,pady=2)
            globals()[f'r_value{i}'].grid(row=i+2,column=2,padx=2,pady=2)
            globals()[f'r_tani{i}'].grid(row=i+2,column=3,padx=2,pady=2)
            globals()[f'r_kingaku{i}'].grid(row=i+2,column=4,padx=2,pady=2)
    #!合計金額の表示
    def kingaku_sum(self):
        font_nonBold = f.Font(size=15)
        font = f.Font( weight="bold", size=15)

        self.display_sum=tk.Label(self.frame_sum,text='0')
        self.display_sum['font']=font
        self.display_sum.pack(side=tk.RIGHT,padx=10,ipadx=20)

        syoukei=tk.Label(self.frame_sum,text='小計 :')
        syoukei['font']=font_nonBold
        syoukei.pack(side=tk.RIGHT,padx=7)

    def sum_process(self):
        sum=0
        for i in range(len(self.rfdf)+self.increase):
            if str(globals()[f'daibunrui{i}'].cget('state'))=='disabled':
                    continue
            kingaku=globals()[f'kingaku{i}'].get()
            if kingaku=='':
                kingaku=0
            sum +=int(kingaku)
        self.display_sum['text']='￥'+str(sum)

    #!購入履歴の表示
    def purchase_history(self,event=None):
        var=self.extractvar.get()
        no=globals()[f'no{var}'].get()
        if no!='':
            result=self.df[self.df['no']==int(no)]
            rs_num=len(result)
            if rs_num!=0:
                result.sort_values('no', ascending=False)
                self.r_label['text']='購入履歴があります！'
                self.r_shmei['text']='商品名：'+result.iloc[0]['syouhinmei']
                
                if len(result)>7:
                    for i in range(7):
                        hizuke=str(result.iloc[i]['hizuke'])
                        if hizuke.find('00:00:00')!=-1:
                            hizuke=re.sub('00:00:00','',hizuke)
                        globals()[f'r_hizuke{i}']['text']=hizuke
                        globals()[f'r_value{i}']['text']=result.iloc[i]['kazu']
                        globals()[f'r_tani{i}']['text']=result.iloc[i]['tani']
                        globals()[f'r_kingaku{i}']['text']=result.iloc[i]['net_value']
                else:
                    for i in range(rs_num):
                        hizuke=str(result.iloc[i]['hizuke'])
                        if hizuke.find('00:00:00')!=-1:
                            hizuke=re.sub('00:00:00','',hizuke)
                        globals()[f'r_hizuke{i}']['text']=hizuke
                        globals()[f'r_value{i}']['text']=result.iloc[i]['kazu']
                        globals()[f'r_tani{i}']['text']=result.iloc[i]['tani']
                        globals()[f'r_kingaku{i}']['text']=result.iloc[i]['net_value']
                        z=7-len(result)

                    for i in range(z):
                        globals()[f'r_hizuke{i+rs_num}']['text']=''
                        globals()[f'r_value{i+rs_num}']['text']=''
                        globals()[f'r_tani{i+rs_num}']['text']=''
                        globals()[f'r_kingaku{i+rs_num}']['text']=''
        else:   
                self.r_label['text']='購入履歴はありません'
                self.r_shmei['text']=''
                for i in range(7):
                    globals()[f'r_hizuke{i}']['text']=''
                    globals()[f'r_value{i}']['text']=''
                    globals()[f'r_tani{i}']['text']=''
                    globals()[f'r_kingaku{i}']['text']=''

    #!抽出前の文字列を表示
    def before_extraction(self,event=None):
        if self.sub_win_txt is None or not self.sub_win_txt .winfo_exists():
            self.sub_win_txt= tk.Toplevel()
            self.sub_win_txt.title("ocr抽出文字列")   # ウィンドウタイトル
            self.sub_win_txt.geometry("550x600") # ウィンドウサイズ(幅x高さ)
            self.textBox=tk.Text(self.sub_win_txt,width = 540, height = 550)
            self.textBox.pack()
        
        s=self.textBox.get(0., tk.END)
        if s!='':
            self.textBox.delete( 0., tk.END )
        if self.txt!='':
            self.textBox.insert(tk.END,self.txt)
    #!抽出前の文字列を表示
    def before_extraction_verMoziretu(self,event=None):
        if self.sub_win_txt is None or not self.sub_win_txt .winfo_exists():
            self.sub_win_txt= tk.Toplevel()
            self.sub_win_txt.title("ocr抽出文字列")   # ウィンドウタイトル
            self.sub_win_txt.geometry("550x600") # ウィンドウサイズ(幅x高さ)
            self.textBox=tk.Text(self.sub_win_txt,width = 540, height = 550)
            self.textBox.pack()
        
        s=self.textBox.get(0., tk.END)
        if s!='':
            self.textBox.delete( 0., tk.END )
        if self.moziretu!='':
            self.textBox.insert(tk.END,self.moziretu)

    #!商品名等の表記
    def insert_display(self,i,index,df,list_dai_bunrui,list_tyuu_bunrui,kazu_list,list_tani):
        globals()[f'puru{i}']=tk.Radiobutton(self.frame_right,value=i,variable=self.extractvar,command=lambda:[self.input_kensaku_entry(),self.extractvar.get()],name=f'puru{i}')
        globals()[f'daibunrui{i}']=ttk.Combobox(self.frame_right,values=list_dai_bunrui,width=12,foreground=self.color,name=f'daibunrui{i}')
        globals()[f'tyubunrui{i}']=ttk.Combobox(self.frame_right,values=list_tyuu_bunrui,width=12,name=f'tyuubunrui{i}')
        globals()[f'no{i}']=tk.Entry(self.frame_right,width=10,name=f'no{i}') 
        globals()[f'syokuhinmei{i}']=tk.Entry(self.frame_right,width=16,name=f'syouhinmei{i}') 
        globals()[f'kazu{i}']=ttk.Combobox(self.frame_right,values=kazu_list,width=6,name=f'kazu{i}')
        globals()[f'tani{i}']=ttk.Combobox(self.frame_right,values=list_tani,width=8,name=f'tani{i}')
        globals()[f'unregistered{i}']=tk.Label(self.frame_right,text='',anchor=tk.E)
        globals()[f'kingaku{i}']=tk.Entry(self.frame_right,width=8,name=f'kingaku{i}')
        globals()[f'memo{i}']=tk.Entry(self.frame_right,width=12,name=f'memo{i}')

        globals()[f'daibunrui{i}'].set(df.iloc[index]["dai_bunrui"])
        globals()[f'tyubunrui{i}'].set(df.iloc[index]["tyuu_bunrui"])
        globals()[f'no{i}'].insert(tk.END,df.iloc[index]["no"])
        globals()[f'syokuhinmei{i}'].insert(tk.END,df.iloc[index]["syouhinmei"])
        globals()[f'kazu{i}'].set(df.iloc[index]["kazu"])
        globals()[f'tani{i}'].set(df.iloc[index]["tani"])
        globals()[f'kingaku{i}'].insert(tk.END,df.iloc[index]["net_value"])
        globals()[f'memo{i}'].insert(tk.END,df.iloc[index]["memo"])
        
        globals()[f'puru{i}'].grid(row=i+4,column=0)
        globals()[f'daibunrui{i}'].grid(row=i+4,column=1,padx=2,pady=2)
        globals()[f'tyubunrui{i}'].grid(row=i+4,column=2,padx=2,pady=2)
        globals()[f'no{i}'].grid(row=i+4,column=3,padx=2,pady=2)
        globals()[f'syokuhinmei{i}'].grid(row=i+4,column=4,padx=2,pady=2)
        globals()[f'kazu{i}'].grid(row=i+4,column=5,padx=2,pady=2)
        globals()[f'tani{i}'].grid(row=i+4,column=6,padx=2,pady=2)
        globals()[f'kingaku{i}'].grid(row=i+4,column=7,padx=2,pady=2)
        globals()[f'memo{i}'].grid(row=i+4,column=8,padx=2,pady=2)
        globals()[f'unregistered{i}'].grid(row=i+4,column=9,padx=2,pady=2)

        self.active_list.append(True)
        self.base_df_list.append(i)
        self.unregistered(i)
        self.nebiki(i)
    

    #!値引き金額をmemo欄に表示
    def nebiki(self,i):
        memo=globals()[f'memo{i}'].get()
        color=str(globals()[f'kingaku{i}'].cget('foreground'))
        #既に金額欄が赤いか
        if color=='red':
            #’％’がない場合（元の金額へ戻す）
            if str(memo)=='' or memo.find('%')==-1: 
                globals()[f'kingaku{i}'].delete( 0, tk.END )
                globals()[f'kingaku{i}'].insert(tk.END,self.rfdf.iloc[i]["net_value"])
                globals()[f'kingaku{i}'].config(foreground="black")
            # '%'がある場合
            else:
                base=self.base_df_list[i]
                globals()[f'memo{i}'].delete( 0, tk.END )
                globals()[f'memo{i}'].insert(tk.END,self.rfdf.iloc[base]["memo"])
        else:
        #memo欄に’％’が記入されていた場合
            if str(memo)!='' and memo.find('%')!=-1:   
                kingaku=int(globals()[f'kingaku{i}'].get())
                nebiki=re.sub(r"\D", "",memo)
                if nebiki!='':
                    hiku=round(kingaku*(int(nebiki)/100))
                    kingaku -=hiku
                    globals()[f'memo{i}'].insert(tk.END,'(-'+str(hiku)+')')
                    globals()[f'kingaku{i}'].delete( 0, tk.END )
                    globals()[f'kingaku{i}'].insert(tk.END,kingaku)
                    globals()[f'kingaku{i}'].config(foreground="red")
        #既に金額欄が赤く、値引き後の金額の場合

    #* -------------------------------------------------------------------------------
    #* 検索窓の作成
    #* -------------------------------------------------------------------------------
    #!検索窓の作成  
    def category_kensaku(self,event=None):
        cate_df=self.category_df
        def kensaku_kekka(tree,event=None):
                tree.delete(*tree.get_children())
                keyword=input_str.get()
                if not keyword:
                    for i in range(len(cate_df)):
                        self.insert_tree(tree,cate_df,i,caller='')
                else:
                    if self.var_kensaku.get()==0:
                        result = cate_df[cate_df['dai_bunrui'].str.contains(keyword,case=False,regex=False)==True]
                    elif self.var_kensaku.get()==1:
                        result = cate_df[cate_df['tyuu_bunrui'].str.contains(keyword,case=False,regex=False)==True]
                    elif self.var_kensaku.get()==2:
                        result = cate_df[cate_df['syokuhin_mei'].str.contains(keyword,case=False,regex=False)==True]
                    if len(result) != 0:
                        for i in range(len(result)):
                            self.insert_tree(tree,result,i,caller='')
        def extract_henkou(event=None):
            selected = tree.focus()
            temp = tree.item(selected,'values')
            var=self.extractvar.get()
            globals()[f'daibunrui{var}'].set(temp[1])
            globals()[f'tyubunrui{var}'].set(temp[2])
            #globals()[f'syokuhinmei{var}'].insert(tk.END,temp[2])
            globals()[f'no{var}'].delete( 0, tk.END )
            globals()[f'no{var}'].insert(tk.END,temp[3])
            globals()[f'tani{var}'].set(temp[6])
            self.unregistered(var)
        self.var_kensaku = tk.IntVar(value=2)

        self.dai_kensaku=tk.Radiobutton(self.frame_raight_bottomLeft,text='大分類',value=0,variable=self.var_kensaku )
        self.tyuu_kensaku=tk.Radiobutton(self.frame_raight_bottomLeft,text='中分類',value=1,variable=self.var_kensaku)
        self.syouhinmei_kensaku=tk.Radiobutton(self.frame_raight_bottomLeft,text='商品名',value=2,variable=self.var_kensaku )
        self.dai_kensaku.grid(row=0,column=0)
        self.tyuu_kensaku.grid(row=0,column=1)
        self.syouhinmei_kensaku.grid(row=0,column=2)
        input_str = tk.StringVar()
        self.kensaku_et=tk.Entry(self.frame_raight_bottomLeft,textvariable=input_str)
        self.kensaku_et.grid(row=1,column=0,columnspan = 2,sticky=tk.EW,padx=5,pady=3)

        tree=self.create_tree_cate_df(self.frame_raight_bottomLeft,caller='')
        tree.bind("<Double-1>", extract_henkou)
        
        for i in range(len(cate_df)):
            self.defo_insert_tree(tree,cate_df,i)

        input_get=tk.Button(self.frame_raight_bottomLeft,text="検索",command=lambda:kensaku_kekka(tree,self))
        input_get.bind_all("<Return>",lambda event:kensaku_kekka(tree,event))
        input_get.grid(row=1,column=2,padx=3,pady=3)

        scrollbar = ttk.Scrollbar(self.frame_raight_bottomLeft, orient=tk.VERTICAL, command=tree.yview)
        scrollbar.grid(row=2,column=7,sticky = tk.N+tk.S)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.grid(row=2,column=0,padx=5,pady=5,columnspan = 7)
    #* -------------------------------------------------------------------------------
    #* categoryへの登録
    #* -------------------------------------------------------------------------------
    #!登録するcategory表の作成
    def create_category_df(self,No,shNo,sh_bunrui,daiNo,dai_bunrui,tyuuNo,tyuu_bunrui,tani,state,syouhinNo,syokuhin_mei):
            result_temp = pd.DataFrame({"No":[No],
                                        "shNo":[shNo],
                                        "sh_bunrui":[sh_bunrui],
                                        "daiNo":[daiNo],
                                        "dai_bunrui":[dai_bunrui],
                                        "tyuuNo":[tyuuNo],
                                        "tyuu_bunrui":[tyuu_bunrui],
                                        "tani":[tani],
                                        "state":[state],
                                        "syouhinNo":[syouhinNo],
                                        "syokuhin_mei":[syokuhin_mei]})
            self.rs_category_df = pd.concat([self.rs_category_df,result_temp])
    #!rs_categoryを初期へ戻す
    def syoki_category(self):
        cols = ['No','shNo','sh_bunrui','daiNo','dai_bunrui','tyuuNo','tyuu_bunrui','tani','state','syouhinNo','syokuhin_mei']
        self.rs_category_df = pd.DataFrame(index=[], columns=cols)
    
    #!rs_category登録前の確認
    def advance_register(self,event=None):
        self.statusbar["text"] = " category register advance_check "
        self.current = 0
        self.delete_list=[]
        self.edit_list=[]
        self.category_register_list={}
        self.syoki_category()

        for i in range(len(self.rfdf)+self.increase):
            if  globals()[f'unregistered{i}'].cget("text")=="!":
                if globals()[f'daibunrui{i}'].get()=="不明" and globals()[f'tyubunrui{i}'].get()=="不明":
                    messagebox.showerror('エラー', "大分類または中分類に’不明’があります")
                    return
                elif globals()[f'tani{i}'].get()=='nan' or globals()[f'tani{i}'].get()=='NaN':
                    messagebox.showerror('エラー', "単位が未入力のものがあります")
                    return
                elif globals()[f'syokuhinmei{i}'].get()=='':
                    messagebox.showerror('エラー', "商品名が未入力のものがあります")
                    return
                No=globals()[f'no{i}'].get()
                same_no_df=self.state_extract(No,callee='ag')
                self.category_register_list[i]=same_no_df
        if  len(self.category_register_list)==0:
            messagebox.showerror('エラー', "登録するcategoryはありません")
            return
        else:
            self.duplication= [k for k, v in self.category_register_list.items() if v == True]
            self.duplication_num=len(self.duplication)
        self.rg_key=iter(self.category_register_list.keys())
        self.category_register(next(self.rg_key))
    
    #!重複番号✔
    def state_extract(self,no,callee):
        if no!='':
            result=self.category_df[self.category_df['No']==int(no)] 
            if len(result)>=1:
                state_ex=result[result['state']=='★']
                if len(state_ex)<1:
                    if callee=='rg':
                        return result
                    else:
                        return True
                else:
                    if callee=='rg':
                        return None
                    else:
                        return False
        else:
            return False
    #!category登録
    def category_register(self,i):
        self.statusbar["text"] = " category register Now "
        self.unregistered(i)
        dai_bunrui=globals()[f'daibunrui{i}'].get()
        tyuu_bunrui=globals()[f'tyubunrui{i}'].get()
        syokuhinmei=globals()[f'syokuhinmei{i}'].get()
        tani=globals()[f'tani{i}'].get()
        #No記入済みであれば、else、未記入はユニーク番号作成
        No=globals()[f'no{i}'].get()
        if No=='':
            #商品名のみの登録（大分類、中分類は引継ぎ
            dai_boolean=str(globals()[f'daibunrui{i}'].cget('foreground'))=='red'
            tyuu_boolean=str(globals()[f'tyubunrui{i}'].cget('foreground'))=='red'
            syouhi_boolean=str(globals()[f'syokuhinmei{i}'].cget('foreground'))=='red'
            if dai_boolean==False and tyuu_boolean==False and syouhi_boolean==True:
                cate_rs=self.category_df[self.category_df['tyuu_bunrui']==tyuu_bunrui]
                shNo=cate_rs.iloc[0]["shNo"]
                daiNo=cate_rs.iloc[0]["daiNo"]
                tyuuNo=cate_rs.iloc[0]["tyuuNo"]
                syouhinNo=cate_rs["syouhinNo"].max()+1
                No=str(shNo)+self.two_fig(daiNo)+self.three_fig(tyuuNo)+self.two_fig(syouhinNo)
            #中分類・商品名を新しく登録する場合
            elif dai_boolean==False and tyuu_boolean==True and syouhi_boolean==True:
                cate_rs=self.category_df[self.category_df['dai_bunrui']==dai_bunrui]
                shNo=cate_rs.iloc[0]["shNo"]
                daiNo=cate_rs.iloc[0]["daiNo"]
                tyuuNo=cate_rs["tyuuNo"].max(axis=0)+1
                syouhinNo=1
                No=str(shNo)+self.two_fig(daiNo)+self.three_fig(tyuuNo)+self.two_fig(syouhinNo)
            #大・中・商品すべて赤文字はエラーへ
            elif dai_boolean==True and tyuu_boolean==True and syouhi_boolean==True:
                self.syoki_category()
                messagebox.showerror('エラー', "入力エラー")
                return
            #作成したNoを入力する
            if str(globals()[f'no{i}'].cget('state'))=='disabled':
                globals()[f'no{i}'].config(state='normal')
            globals()[f'no{i}'].insert(tk.END,No)
        else:
            shNo=int(No[0])
            daiNo=int(No[1:3])
            tyuuNo=int(No[3:6])
            syouhinNo=int(No[6:])
        sh_bunrui=self.bunrui_check(shNo)
        self.state=''             
        self.current += 1
    
        if self.category_register_list[i]==True:
            if self.sub_win_noCheck is None or not self.sub_win_noCheck.winfo_exists():
                self.sub_win_noCheck= tk.Toplevel()
                self.sub_win_noCheck.title("重複したNoがあります"+'('+str(self.duplication.index(i)+1)+'/'+str(self.duplication_num)+')')   # ウィンドウタイトル
                self.sub_win_noCheck.geometry("550x200")
                self.sub_win_noCheck.focus_set()
            label_main=tk.Label(self.sub_win_noCheck,text='他名称の登録:',bg = 'cyan1')
            label_main.grid(row=1,column=0,pady=5)
            self.rg_syouhinmei=tk.Entry(self.sub_win_noCheck) 
            self.rg_syouhinmei.grid(row=1,column=1,columnspan=2,pady=5,sticky=tk.EW)

            self.main_bln = tk.BooleanVar()
            self.main_bln.set(True)
            chk = tk.Checkbutton(self.sub_win_noCheck, text='mainに登録',variable=self.main_bln)
            chk.grid(row=1,column=4,pady=5)

            same_no_df=self.state_extract(No,'rg')
            tree=self.create_tree_cate_df(self.sub_win_noCheck,caller='cate_rs')
            same_no_df_num=len(same_no_df)
            tree.configure(height=int(same_no_df_num)+1)
            tree.insert("", "end", values=('追加',
                                            sh_bunrui,
                                            dai_bunrui, 
                                            tyuu_bunrui,
                                            No,
                                            self.state,
                                            syokuhinmei,
                                            tani))
            for i in range(len(same_no_df)):
                self.insert_tree(tree,same_no_df,i,caller='cate_rs')
            tree.grid(row=2,column=0,columnspan=6,padx=10,pady=5,sticky=tk.EW)
            tree.bind("<Double-1>", lambda event:self.state_henkou(tree,event))

            register_button=tk.Button(self.sub_win_noCheck,text="登録",
                                      command=lambda:self.formal_name_register(tree,same_no_df_num,No,shNo,sh_bunrui,daiNo,dai_bunrui,tyuuNo,tyuu_bunrui,tani,syouhinNo))
            register_button.grid(row=1,column=3,pady=5)

            if self.current  < len(self.category_register_list):
                input_get=tk.Button(self.sub_win_noCheck,text="次へ",
                                    command=lambda:[self.sub_win_noCheck.destroy(),
                                                    self.create_category_df(No,shNo,sh_bunrui,daiNo,dai_bunrui,tyuuNo,tyuu_bunrui,tani,self.state,syouhinNo,syokuhinmei),
                                                    self.category_register(next(self.rg_key))])
                input_get.grid(row=3,column=5,padx=5,pady=5)
            else:
                input_get=tk.Button(self.sub_win_noCheck,text="category登録へ",
                                    command=lambda:[self.sub_win_noCheck.destroy(),
                                                    self.create_category_df(No,shNo,sh_bunrui,daiNo,dai_bunrui,tyuuNo,tyuu_bunrui,tani,self.state,syouhinNo,syokuhinmei),
                                                    self.confi_rs_category()])
                input_get.grid(row=3,column=2,padx=5,pady=5)
        else:
            self.create_category_df(No,shNo,sh_bunrui,daiNo,dai_bunrui,tyuuNo,tyuu_bunrui,tani,self.state,syouhinNo,syokuhinmei)
            if self.current  < len(self.category_register_list):
                self.category_register(next(self.rg_key))
            else:
                self.confi_rs_category()
        
    #!treeを選択したときの動作
    def state_henkou(self,tree,event=None):
        answer = messagebox.askokcancel('確認', 'これをメインにしますか？')
        if answer==True:
            selected = tree.focus()
            temp = tree.item(selected,'values')
            if temp[0]=='追加':
                self.state='★'
                tree.item(selected,value=(temp[0],temp[1],temp[2],temp[3],temp[4],self.state,temp[6],temp[7]))
            else:
                if temp[0]=='削除':
                    messagebox.showerror('エラー','削除予定の行です')
                    self.sub_win_noCheck.focus_set()
                    return
                self.index=list(self.category_df.reset_index().query('syokuhin_mei == @temp[6]').index)[0]
                self.edit_list.append(self.index+2)
                tree.item(selected,value=(temp[0],temp[1],temp[2],temp[3],temp[4],'★',temp[6],temp[7]))
        self.sub_win_noCheck.focus_set()
    #!新しい名称登録 
    def formal_name_register(self,tree,same_no_df_num,No,shNo,sh_bunrui,daiNo,dai_bunrui,tyuuNo,tyuu_bunrui,tani,syouhinNo):
        syokuhinmei=self.rg_syouhinmei.get()
        answer =messagebox.askokcancel(title='確認', message='名称を新しく登録しますか',detail='商品名:'+syokuhinmei)
        if answer==True:
            state=''
            if self.main_bln.get()==True:
                state='★'
            self.create_category_df(No,shNo,sh_bunrui,daiNo,dai_bunrui,tyuuNo,tyuu_bunrui,tani,state,syouhinNo,syokuhinmei)
            tree.insert("", "end", values=('追加',
                                            sh_bunrui,
                                            dai_bunrui, 
                                            tyuu_bunrui,
                                            No,
                                            state,
                                            syokuhinmei,
                                            tani))
            tree.configure(height=int(same_no_df_num)+2)
            messagebox.showinfo("メッセージ", "登録完了")
            self.sub_win_noCheck.focus_set()
        else:
            self.sub_win_noCheck.focus_set()
            return

    #!登録するcategoryをモーダルウィンドウで確認
    def confi_rs_category(self):
        if self.sub_win is None or not self.sub_win.winfo_exists():
            if not(self.sub_win is None or not self.sub_win.winfo_exists()):
                self.sub_win_noCheck.destroy()
            self.sub_win= tk.Toplevel()
            self.sub_win.title("categoryへ登録確認画面")   # ウィンドウタイトル
            self.sub_win.geometry("550x200") # ウィンドウサイズ(幅x高さ)

            #Excelへの登録
            def register_category(event=None):    
                try:
                    rr=openpyxl.load_workbook(self.excel_file_path)
                    category_sheet=rr['category']
                    #categoryの登録
                    rows= dataframe_to_rows(self.rs_category_df, index=False, header=False)
                    row_start_index=category_sheet.max_row+1
                    col_start_index=1
                    for row_no, row in enumerate(rows,row_start_index):
                        for col_no, value in enumerate(row,col_start_index):
                            category_sheet.cell(row=row_no, column=col_no, value=value) # 1セルづつ書込む
                    #categoryの変更
                    if len(self.edit_list)!=0:
                        for e in self.edit_list:
                            category_sheet.cell(row=e, column=9, value='★')
                    #categoryの削除
                    if len(self.delete_list)!=0:
                        for d in self.delete_list:
                            category_sheet.delete_rows(d)
                    rr.save(self.excel_file_path)
                    messagebox.showinfo("メッセージ", "登録完了")
                    #登録後モーダルウィンドウを閉じて、登録後のExcel表をアップロード、検索窓も更新
                    self.sub_win.destroy()
                    self.rireki_excel_get()
                    self.category_kensaku()
                    self.update_unregistered()
                    messagebox.showinfo("メッセージ", "categoryを更新しました")
                except PermissionError:
                    answer=messagebox.askyesno('Excelファイルが開かれています','Excelを閉じて(自動保存）、もう一度登録しますか？)')
                    if answer==True:
                        excel = win32com.client.Dispatch("Excel.Application")
                        book = excel.Workbooks.Open(self.excel_file_path)
                        book.Save()
                        book.Close()
                        excel.Quit()
                        register_category()
                    else:
                        self.syoki_category()
                        self.sub_win.destroy()
                        pass
                except Exception as e:
                    self.syoki_category()
                    messagebox.showerror('エラー',e)
                    self.sub_win.focus_set()
            tree=self.create_tree_cate_df(self.sub_win,caller='cate_rs')
            for i in range(len(self.rs_category_df)):
                self.insert_tree(tree,self.rs_category_df,i,caller='cate_rg')
            if len(self.edit_list)!=0:
                for i in self.edit_list:
                    index=i-2
                    state=self.category_df['state'][index]
                    if str(state)=='nan':
                        state=''
                    tree.insert("", "end", values=('編集',
                                                self.category_df['sh_bunrui'][index],
                                                self.category_df['dai_bunrui'][index], 
                                                self.category_df['tyuu_bunrui'][index],
                                                self.category_df['No'][index],
                                                '★',
                                                self.category_df['syokuhin_mei'][index],
                                                self.category_df['tani'][index]),
                                                tags="green")
                    tree.tag_configure("green", foreground="green")
            if len(self.delete_list)!=0:
                for i in self.delete_list:
                    index=i-2
                    state=self.category_df['state'][index]
                    if str(state)=='nan':
                        state=''
                    tree.insert("", "end", values=('削除',
                                                self.category_df['sh_bunrui'][index],
                                                self.category_df['dai_bunrui'][index], 
                                                self.category_df['tyuu_bunrui'][index],
                                                self.category_df['No'][index],
                                                state,
                                                self.category_df['syokuhin_mei'][index],
                                                self.category_df['tani'][index]),
                                                tags="red")
                    tree.tag_configure("red", foreground='red')
            tree.pack()
            tree.bind("<Double-1>", lambda event:self.defo_name_rg(tree,event))

            input_get=tk.Button(self.sub_win,text="登録・変更・削除",command=register_category)
            input_get.pack(pady=5)
    
    # !最終確認画面でも新しい名称を登録できるようにする
    def defo_name_rg(self,tree,event=None):
        selected = tree.focus()
        temp = tree.item(selected,'values')
        i=list(self.rs_category_df.reset_index().query('syokuhin_mei == @temp[6]').index)[0]
        if i=='':
            messagebox.showerror('エラー', "既に登録済みで削除予定または変更予定です。選択できません")
            self.sub_win.focus_set()
            return
        self.sub_win.destroy()
        if self.sub_win_noCheck is None or not self.sub_win_noCheck.winfo_exists():
                self.sub_win_noCheck= tk.Toplevel()
                self.sub_win_noCheck.title("他名称の登録")   # ウィンドウタイトル
                self.sub_win_noCheck.geometry("550x200")
                self.sub_win_noCheck.focus_set()
   
        label_main=tk.Label(self.sub_win_noCheck,text='他名称の登録:',bg = 'cyan1')
        label_main.grid(row=1,column=0,pady=5)
        self.rg_syouhinmei=tk.Entry(self.sub_win_noCheck) 
        self.rg_syouhinmei.grid(row=1,column=1,columnspan=2,pady=5,sticky=tk.EW)

        self.main_bln = tk.BooleanVar()
        self.main_bln.set(True)
        chk = tk.Checkbutton(self.sub_win_noCheck, text='mainに登録',variable=self.main_bln)
        chk.grid(row=1,column=4,pady=5)
        
        tree2=self.create_tree_cate_df(self.sub_win_noCheck,caller='cate_rs')
        self.insert_tree(tree2,self.rs_category_df,i,caller='cate_rs')
        tree2.configure(height=1)

        tree2.grid(row=2,column=0,columnspan=6,padx=10,pady=5,sticky=tk.EW)
        tree2.bind("<Double-1>", lambda event:self.state_henkou(tree2,event))

        same_no_df_num=1
        No=self.rs_category_df['No'].iloc[i]
        shNo=self.rs_category_df['shNo'].iloc[i]
        sh_bunrui=self.rs_category_df['sh_bunrui'].iloc[i]
        daiNo=self.rs_category_df['daiNo'].iloc[i]
        dai_bunrui=self.rs_category_df['dai_bunrui'].iloc[i]
        tyuu_bunrui=self.rs_category_df['tyuu_bunrui'].iloc[i]
        tyuuNo=self.rs_category_df['tyuuNo'].iloc[i]
        tani=self.rs_category_df['tani'].iloc[i]
        syouhinNo=self.rs_category_df['syouhinNo'].iloc[i]
        self.rs_category_df['tani'].iloc[i]
        register_button=tk.Button(self.sub_win_noCheck,text="登録",
                                    command=lambda:self.formal_name_register(tree2,same_no_df_num,No,shNo,sh_bunrui,daiNo,dai_bunrui,tyuuNo,tyuu_bunrui,tani,syouhinNo))
        register_button.grid(row=1,column=3,pady=5)
        back_button=tk.Button(self.sub_win_noCheck,text="戻る",command=lambda:[self.confi_rs_category(),self.sub_win_noCheck.destroy()])
        back_button.grid(row=3,column=2,pady=5)
        
    #* -------------------------------------------------------------------------------
    #* rirekiへの登録
    #* -------------------------------------------------------------------------------
    #!初期値の設定
    def syoki_df(self):
        cols = ['kirokuNo','tempo', 'hizuke','dai_bunrui','tyuu_bunrui','no','syouhinmei','kazu','tani','net_value','memo']
        self.rs_result_df = pd.DataFrame(index=[], columns=cols)
    #!登録するrs_df表の作成
    def create_rs_result_df(self,kirokuNo,tempo,hizuke,dai_bunrui,tyuu_bunrui,no,syouhinmei,kazu,tani,kingaku,memo):
            result_temp = pd.DataFrame({"kirokuNo":[kirokuNo],
                                        "tempo":[tempo],
                                        "hizuke":[hizuke],
                                        "dai_bunrui":[dai_bunrui],
                                        "tyuu_bunrui":[tyuu_bunrui],
                                        "no":[no],
                                        "syouhinmei":[syouhinmei],
                                        "kazu":[kazu],
                                        "tani":[tani],
                                        "net_value":[kingaku],
                                        "memo":[memo]})
            self.rs_result_df = pd.concat([self.rs_result_df,result_temp])

    def check_dis_register(self,event=None):
        disabled_num=0
        for i in range(len(self.rfdf)+self.increase):
            if str(globals()[f'daibunrui{i}'].cget('state'))=='disabled':
                    disabled_num +=1
                    continue
            else:
                tempo=self.tempo_combobox.get()
                dai_boolean=str(globals()[f'daibunrui{i}'].cget('foreground'))=='red'
                tyuu_boolean=str(globals()[f'tyubunrui{i}'].cget('foreground'))=='red'
                syouhi_boolean=str(globals()[f'syokuhinmei{i}'].cget('foreground'))=='red'
                if tempo=='' or tempo=='不明':
                    messagebox.showerror('エラー', "店舗名が未記入です")
                    return
                hizuke=self.hizuke.get()
                if hizuke=='' or hizuke=='不明':
                    messagebox.showerror('エラー', "日付が未記入です")
                    return
                daibunnrui=globals()[f'daibunrui{i}'].get()
                tyuu_bunrui=globals()[f'tyubunrui{i}'].get()
                if daibunnrui=='不明' or daibunnrui=='' or tyuu_bunrui=='不明':
                    messagebox.showerror('エラー', "未分類の商品があります")
                    return
                no=globals()[f'no{i}'].get()
                if no=='':
                    messagebox.showerror('エラー', "noが未記入のものがあります")
                    return
                syouhinmei=globals()[f'syokuhinmei{i}'].get()
                if syouhinmei=='':
                    messagebox.showerror('エラー', "商品名が未記入ものがあります")
                    return
                kingaku=globals()[f'kingaku{i}'].get()
                if kingaku=='':
                    messagebox.showerror('エラー', "金額が未記入ものがあります")
                    return
                if dai_boolean==True or tyuu_boolean==True or syouhi_boolean==True:
                    messagebox.showerror('エラー', "category未登録のものがあります")
                    return
        if len(self.rfdf)+self.increase<=disabled_num:
            messagebox.showerror('エラー', "登録するものがありません")
            return
        self.rs_rs_df_create()

    #!修正後のrs表を作成
    def rs_rs_df_create(self,event=None):
        kirokuNo=self.df.iloc[-1]['kirokuNo']
        try:
            for i in range(len(self.rfdf)+self.increase):
                if str(globals()[f'daibunrui{i}'].cget('state'))=='disabled':
                    continue
                else:
                    kirokuNo=kirokuNo+1
                    tempo=self.tempo_combobox.get()
                    hizuke=self.hizuke.get()

                    dai_bunrui=globals()[f'daibunrui{i}'].get()
                    tyuu_bunrui=globals()[f'tyubunrui{i}'].get()
                    no=globals()[f'no{i}'].get()
                    syouhinmei=globals()[f'syokuhinmei{i}'].get()
                    kazu=globals()[f'kazu{i}'].get()
                    tani=globals()[f'tani{i}'].get()
                    kingaku=globals()[f'kingaku{i}'].get()
                    memo=globals()[f'memo{i}'].get()
                self.create_rs_result_df(kirokuNo,tempo,hizuke,dai_bunrui,tyuu_bunrui,no,syouhinmei,kazu,tani,kingaku,memo)
            self.register_rs_result_df(hizuke,tempo)
        except Exception as e:
            self.syoki_df()
            messagebox.showerror('エラー',e)
    
    #!Excelに保存
    def register_rs_result_df(self,hizuke,tempo):
        if os.path.exists(self.filename)==False:
            messagebox.showerror('エラー','既に登録済みのレシートです')
            return
        try:
            rr=openpyxl.load_workbook(self.excel_file_path)
            rireki_sheet=rr['rireki']
            rows= dataframe_to_rows(self.rs_result_df,index=False, header=False)
            row_start_index=rireki_sheet.max_row+1
            col_start_index=1
            for row_no, row in enumerate(rows,row_start_index):
                for col_no, value in enumerate(row,col_start_index):
                    rireki_sheet.cell(row=row_no, column=col_no, value=value) # 1セルづつ書込む
            rr.save(self.excel_file_path)
            messagebox.showinfo("メッセージ", "登録完了")
            basename=self.filename[:self.filename.rfind('\\')+1]
            hizuke=re.sub(r'\D','',hizuke)
            rename_filename=basename+hizuke+'_'+tempo+'.jpg'
            os.rename(self.filename,rename_filename)
            shutil.move(rename_filename,r"C:\Users\f1351\Desktop\レシート\済：レシート")
            messagebox.showinfo("メッセージ", "ファイル名を変更して、完了フォルダー移動しました")
        except PermissionError:
            answer=messagebox.askyesno('Excelファイルが開かれています','Excelを閉じて(自動保存）、もう一度登録しますか？)')
            if answer==True:
                excel = win32com.client.Dispatch("Excel.Application")
                book = excel.Workbooks.Open(self.excel_file_path)
                book.Save()
                book.Close()
                excel.Quit()
                self.register_rs_result_df(hizuke,tempo)
            else:
                self.syoki_df()
                self.sub_win.destroy()
                pass
        except Exception as e:
            shutil.move(rename_filename,r"C:\Users\f1351\Desktop")
            self.syoki_df()
            messagebox.showerror('エラー',e)

    #* -------------------------------------------------------------------------------
    #*Excelファイル読み込み、初期値設定
    #* -------------------------------------------------------------------------------
    # !
    def open_excel(self,event=None):
        excel = win32com.client.Dispatch("Excel.Application")
        book = excel.Workbooks.Open(self.excel_file_path)
        excel.Visible = True
    #!初期rfdeまた登録rs_category_dfの初期値登録
    def syokiRFDF(self):
        self.syoki_category()
        self.syoki_df()
        result_temp = pd.DataFrame({"kirokuNo":[''],
                                "tempo":[''],
                                "hizuke":[''],
                                "dai_bunrui":[''],
                                "tyuu_bunrui":[''],
                                "no":[''],
                                "syouhinmei":[''],
                                "kazu":[''],
                                "tani":[''],
                                "net_value":[''],
                                "memo":['']})
        return result_temp

    #!Excelの抽出
    def rireki_excel_get(self,event=None):
        try:
            self.df=pd.read_excel(self.excel_file_path,sheet_name='rireki',index_col=None)
            self.category_df=pd.read_excel(self.excel_file_path,sheet_name='category',index_col=None)
            self.tempo_df=pd.read_excel(self.excel_file_path,sheet_name='tempo',index_col=None)           
        except Exception as e:
            print(e)
    #* -------------------------------------------------------------------------------
    #*そのほか　表記変更　ユニーク番号、数字2桁
    #* -------------------------------------------------------------------------------
    #!登録するcategoryをtreeで表示 
    def create_tree_cate_df(self,frame,caller):
        if caller=='cate_rs':
            tree = ttk.Treeview(frame, columns=["EC","分類","大分類", "中分類","No","s","商品名","単位"],height=6,show='headings',selectmode='browse')
            tree.column("EC", width=50)
            tree.heading("EC", text='EC')
        else:    
            tree = ttk.Treeview(frame, columns=["分類","大分類", "中分類","No","s","商品名","単位"],height=6,show='headings',selectmode='browse')
        #tree.bind("<Double-1>", extract_henkou)
        tree.column("分類", width=50)
        tree.column("大分類", width=70)
        tree.column("中分類", width=100)
        tree.column("No", width=70)
        tree.column("s", width=25)
        tree.column("商品名", width=100)
        tree.column("単位", width=50)

        tree.heading("分類", text='分類')
        tree.heading("大分類", text='大分類')
        tree.heading("中分類", text='中分類')
        tree.heading("No", text="No")
        tree.heading("s", text="s")
        tree.heading("商品名", text='商品名',anchor='center')
        tree.heading("単位", text='単位')
        return tree

    #!検索窓　および　category登録の際の表示
    def defo_insert_tree(self,tree,df,i):
        state=df['state'][i]
        if str(state)=='nan':
            state=''
        tree.insert("", "end", values=(df['sh_bunrui'][i],
                                    df['dai_bunrui'][i], 
                                    df['tyuu_bunrui'][i],
                                    df['No'][i],
                                    state,
                                    df['syokuhin_mei'][i],
                                    df['tani'][i]))
    def insert_tree(self,tree,df,i,caller):
        state=df['state'].iloc[i]
        no=df['No'].iloc[i]
        color='black'
        if str(state)=='nan':
            state=''
        if caller=='cate_rs' or caller=='cate_rg' :
            if caller=='cate_rs':   
                if str(df['syokuhin_mei'].iloc[i])=='nan':
                    editedContent='削除'
                    self.index=list(self.category_df.reset_index().query('No == @no').index)[0]
                    self.delete_list.append(self.index+2)
                    color='red'
                else:
                    editedContent=''
            elif caller=='cate_rg':
                editedContent='登録'
            tree.insert("", "end", values=(editedContent,
                                        df['sh_bunrui'].iloc[i],
                                        df['dai_bunrui'].iloc[i], 
                                        df['tyuu_bunrui'].iloc[i],
                                        no,
                                        state,
                                        df['syokuhin_mei'].iloc[i],
                                        df['tani'].iloc[i]),
                                        tags=color)
            tree.tag_configure("red", foreground='red')
            tree.tag_configure("black", foreground='black')
        elif caller=='':
            tree.insert("", "end", values=(df['sh_bunrui'].iloc[i],
                                        df['dai_bunrui'].iloc[i], 
                                        df['tyuu_bunrui'].iloc[i],
                                        no,
                                        state,
                                        df['syokuhin_mei'].iloc[i],
                                        df['tani'].iloc[i]))
    
    #!未登録　確認
    def unregistered(self,i):
        cate_df=self.category_df
        font_bold=f.Font(weight="bold")
        dai=cate_df[cate_df["dai_bunrui"]==globals()[f'daibunrui{i}'].get()]
        if str(globals()[f'daibunrui{i}'].cget('state'))=='disabled':
            return
        if len(dai)==0:
            self.dai_color,self.tyuu_color,self.syouhin_color="red","red","red"
            globals()[f'unregistered{i}'].config(text='!',fg="red",font=font_bold)
        else:
            self.dai_color="black"
            tyuu=cate_df[cate_df["tyuu_bunrui"]==globals()[f'tyubunrui{i}'].get()]
            if len(tyuu)==0:
                self.tyuu_color,self.syouhin_color="red","red"
                globals()[f'unregistered{i}'].config(text='!',fg="red",font=font_bold)
                globals()[f'no{i}'].delete( 0, tk.END )
                globals()[f'no{i}'].config(state='disabled')
            else:
                self.tyuu_color="black"
                syouhin=cate_df[cate_df["syokuhin_mei"]==globals()[f'syokuhinmei{i}'].get()]
                if len(syouhin)==0:
                    self.syouhin_color="red"
                    globals()[f'unregistered{i}'].config(text='!',fg="red",font=font_bold)
                else:
                    self.syouhin_color="black"
                    globals()[f'unregistered{i}']["text"]=""
                    globals()[f'no{i}'].config(state='disabled')

        globals()[f'daibunrui{i}'].config(foreground=self.dai_color)
        globals()[f'tyubunrui{i}'].config(foreground=self.tyuu_color)
        globals()[f'syokuhinmei{i}'].config(fg=self.syouhin_color)
    
    #!食品か日用品か判断
    def bunrui_check(self,no):
        sh_bunrui=''
        if str(no)=='1':
            sh_bunrui="食品"
        elif str(no)=='2':
            sh_bunrui="日用品"
        elif str(no)=='3':
            sh_bunrui="美容品"
        elif str(no)=='4':
             sh_bunrui="医薬品"
        return sh_bunrui
    
    #!ユニークNoの作成
    def create_unique(self,df,i):
        shNo=str(df['shNo'].iloc[i])
        daiNo=self.two_fig(df['daiNo'].iloc[i])
        tyuuNo=self.three_fig(df['tyuuNo'].iloc[i])
        syouhinNo=self.two_fig(df['syouhinNo'].iloc[i])
        No=shNo+daiNo+tyuuNo+syouhinNo
        return No
    def two_fig(self,num):
            two_num=''
            num=str(num)
            if len(num)<2:
                two_num='0'+num
            else:
                two_num=num
            return two_num
    def three_fig(self,num):
        three_num=''
        num=str(num)
        if len(num)<2:
            three_num='0'+'0'+num
        elif len(num)<3:
            three_num='0'+num
        else:
            three_num=num
        return three_num
    #* -------------------------------------------------------------------------------
    #* ショートカットまたはマウス操作
    #* -------------------------------------------------------------------------------
    #!キーボードでのラジオボタン選択変更
    def key_push(self,event):
        if event.keysym == "Up" or event.keysym == "Down":
            if event.keysym == "Up":
                var=self.up_down_check(ud='d')
            elif event.keysym == "Down":
                var=self.up_down_check(ud='u')
            globals()[f'puru{var}'].select()
            globals()[f'syokuhinmei{var}'].focus_set()
            self.input_kensaku_entry()
            self.purchase_history()
        elif event.keysym=="Next" or event.keysym=="Prior":
            if event.keysym=="Next":
                t=self.up_down_check(ud='u')
                # t=1
            elif event.keysym=="Prior":
                t=self.up_down_check(ud='d')
                # t=-1
            self.key_next_prior(t)
            globals()[f'puru{t}'].select()
            globals()[f'syokuhinmei{t}'].focus_set()
            self.input_kensaku_entry()
            self.purchase_history()
        else:
            self.input_kensaku_entry_noinput(event.widget)
    
    #!選択先がdeleteされていないか✔
    def up_down_check(self,ud):
        c=self.extractvar.get()
        if ud=='d':
            var=c-1
            if var<0:
                    var=len(self.active_list)-1
            while self.active_list[var]==False:
                var -=1
                if var<0:
                    var=len(self.active_list)-1
        elif ud=='u':
            var=c+1
            if var>=len(self.active_list):
                    var=0
            while self.active_list[var]==False:
                var +=1
                if var>=len(self.active_list):
                    var=0
        self.base_df_list[c],self.base_df_list[var]=self.base_df_list[var],self.base_df_list[c]
        return var
    #! 前後入れ替えﾌﾟﾛｾｽ
    def key_next_prior(self,t):
        var=self.extractvar.get()
        daibunrui,tyuubunrui,no,syouhinmei,kazu,tani,kingaku,memo,st=self.current_index_get(var)
        n_daibunrui,n_tyuubunrui,n_no,n_syouhinmei,n_kazu,n_tani,n_kingaku,n_memo,n_st=self.current_index_get(t)
        self.henkou_index(t,daibunrui,tyuubunrui,no,syouhinmei,kazu,tani,kingaku,memo,st)
        self.henkou_index(var,n_daibunrui,n_tyuubunrui,n_no,n_syouhinmei,n_kazu,n_tani,n_kingaku,n_memo,n_st)
    
    #!result_dfへの追加
    def tuika_index(self,event=None):
        list_tempo,list_dai_bunrui,list_tyuu_bunrui,list_tani=self.tempo_category_list()
        kazu_list=[1,2,3,4,5,6,7,8,9,10]
        var=len(self.rfdf)+self.increase
        self.increase+=1
        self.base_df_list.append
        result_temp=self.syokiRFDF()
        self.insert_display(var,0,result_temp,list_dai_bunrui,list_tyuu_bunrui,kazu_list,list_tani)
        current=self.extractvar.get()
        if current+1<var:
            t=current+1
            self.base_df_list.insert(t,'None')
            daibunrui,tyuubunrui,no,syouhinmei,kazu,tani,kingaku,memo,st=self.current_index_get(t)
            self.display_index(t+1,daibunrui,tyuubunrui,no,syouhinmei,kazu,tani,kingaku,memo,st,caller="nm")
     #!反映させる
    def display_index(self,t,daibunrui,tyuubunrui,no,syouhinmei,kazu,tani,kingaku,memo,st,caller):
        n_daibunrui,n_tyuubunrui,n_no,n_syouhinmei,n_kazu,n_tani,n_kingaku,n_memo,n_st=self.current_index_get(t)
        self.henkou_index(t,daibunrui,tyuubunrui,no,syouhinmei,kazu,tani,kingaku,memo,st)
        max=len(self.rfdf)+self.increase-1
        if max>t and caller!="fn":
            self.display_index(t+1,n_daibunrui,n_tyuubunrui,n_no,n_syouhinmei,n_kazu,n_tani,n_kingaku,n_memo,n_st,caller="nm")
        elif caller!="fn":
            self.display_index(self.extractvar.get()+1,'','','','','','','','','',caller="fn")
        else:
            pass

     #! 前後入れ替えﾌﾟﾛｾｽ
    def current_index_get(self,ct):
        daibunrui=globals()[f'daibunrui{ct}'].get()
        tyuubunrui=globals()[f'tyubunrui{ct}'].get()
        no=globals()[f'no{ct}'].get()
        syouhinmei=globals()[f'syokuhinmei{ct}'].get()
        kazu=globals()[f'kazu{ct}'].get()
        tani=globals()[f'tani{ct}'].get()
        kingaku=globals()[f'kingaku{ct}'].get()
        memo=globals()[f'memo{ct}'].get()
        st=''
        if str(globals()[f'daibunrui{ct}'].cget('state'))=='disabled':
            st='disabled'
        return daibunrui,tyuubunrui,no,syouhinmei,kazu,tani,kingaku,memo,st

    #!
    def henkou_index(self,t,daibunrui,tyuubunrui,no,syouhinmei,kazu,tani,kingaku,memo,st):
        # 移動先が削除されていて'disabled'になっていないか✔➡　normalへ
        if str(globals()[f'puru{t}'].cget('state'))=='disabled':
            globals()[f'puru{t}'].config(state='normal')
        if str(globals()[f'daibunrui{t}'].cget('state'))=='disabled':
            globals()[f'daibunrui{t}'].config(state='normal')
        globals()[f'daibunrui{t}'].set(daibunrui)
        if str(globals()[f'tyubunrui{t}'].cget('state'))=='disabled':
            globals()[f'tyubunrui{t}'].config(state='normal')
        globals()[f'tyubunrui{t}'].set(tyuubunrui)
        if str(globals()[f'no{t}'].cget('state'))=='disabled':
            globals()[f'no{t}'].config(state='normal')
        globals()[f'no{t}'].delete( 0, tk.END )
        globals()[f'no{t}'].insert(tk.END,no)
        if str(globals()[f'syokuhinmei{t}'].cget('state'))=='disabled':
            globals()[f'syokuhinmei{t}'].config(state='normal')
        globals()[f'syokuhinmei{t}'].delete( 0, tk.END )
        globals()[f'syokuhinmei{t}'].insert(tk.END,syouhinmei)
        if str(globals()[f'kazu{t}'].cget('state'))=='disabled':
            globals()[f'kazu{t}'].config(state='normal')
        globals()[f'kazu{t}'].set(kazu)
        if str(globals()[f'tani{t}'].cget('state'))=='disabled':
            globals()[f'tani{t}'].config(state='normal')
        globals()[f'tani{t}'].set(tani)
        if str(globals()[f'kingaku{t}'].cget('state'))=='disabled':
            globals()[f'kingaku{t}'].config(state='normal')
        globals()[f'kingaku{t}'].delete( 0, tk.END )
        globals()[f'kingaku{t}'].insert(tk.END,kingaku)
        if str(globals()[f'memo{t}'].cget('state'))=='disabled':
            globals()[f'memo{t}'].config(state='normal')
        globals()[f'memo{t}'].delete( 0, tk.END )
        globals()[f'memo{t}'].insert(tk.END,memo)
        if memo!='' and memo.find('%')!=-1:
            globals()[f'kingaku{t}'].config(foreground="red")
        if str(globals()[f'kingaku{t}'].cget('foreground'))=='red':
            if memo=='' or memo.find('%')==-1:   
                globals()[f'kingaku{t}'].config(foreground="black")
        self.unregistered(t)
        if st=='disabled':
            self.delete_row(t)

    #!検索窓へフォーカスさせる
    def kensaku_entry_focus(self,event):
        sss='!canvas2.!frame.'
        if sss in str(event.widget):
            self.kensaku_et.focus_set()
        else:
            var=self.extractvar.get()
            globals()[f'syokuhinmei{var}'].focus_set()

    #!検索対象変更ショートカット
    def key_left(self,event):
        mmm='.!frame2'
        if (mmm in str(event.widget))==False:
            return
        var_kensaku=self.var_kensaku.get()-1
        if var_kensaku<0:
            var_kensaku=2
        self.left_right_check(var_kensaku)
    def key_right(self,event):
        mmm='.!frame2'
        if (mmm in str(event.widget))==False:
            return
        var_kensaku=self.var_kensaku.get()+1
        if var_kensaku>2:
            var_kensaku=0
        self.left_right_check(var_kensaku)

    def left_right_check(self,var_kensaku):
        if var_kensaku==0:
            self.dai_kensaku.select()
        elif var_kensaku==1:
            self.tyuu_kensaku.select()
        elif var_kensaku==2:
            self.syouhinmei_kensaku.select()

    #!クリックしたら、プルダウンも変更
    def entry_click(self,event):
        sss='!canvas2.!frame.'
        if (sss in str(event.widget))==False:
            return
        index=str(event.widget).split(".")[-1]
        index=re.sub(r'\D','',index)
        if index=='':
            return
        globals()[f'puru{index}'].select()
        self.input_kensaku_entry()   
        self.purchase_history(index)

    #ダブルクリックしてdisablesを解除
    def no_double(self,event):
        no_dis='!canvas2.!frame.no'
        puru_dis='!canvas2.!frame.puru'
        index=str(event.widget).split(".")[-1]
        index=re.sub(r'\D','',index)
        if index=='':
            return
        if (no_dis in str(event.widget))==True:
            if str(globals()[f'no{index}'].cget('state'))!='disabled':
                return
            globals()[f'no{index}'].config(state='normal')
        elif (puru_dis in str(event.widget))==True:
            if str(globals()[f'puru{index}'].cget('state'))!='disabled':
                return
            globals()[f'puru{index}'].config(state='normal')
            globals()[f'daibunrui{index}'].config(state='normal')
            globals()[f'tyubunrui{index}'].config(state='normal')
            globals()[f'syokuhinmei{index}'].config(state='normal')
            globals()[f'no{index}'].config(state='normal')
            globals()[f'kazu{index}'].config(state='normal')
            globals()[f'tani{index}'].config(state='normal')
            globals()[f'kingaku{index}'].config(state='normal')
            globals()[f'memo{index}'].config(state='normal')
            self.unregistered(index)
            self.active_list[int(index)]=True

    #!もとに戻すボタン
    def result_initial_value(self,event=None):
        var=self.extractvar.get()
        base=self.base_df_list[var]
        if base=='None':
            messagebox.showerror('エラー','追加された行のため、元の値はありません')
            return
        globals()[f'syokuhinmei{var}'].delete( 0, tk.END )
        globals()[f'syokuhinmei{var}'].insert(tk.END,self.rfdf.iloc[base]["syouhinmei"])

    #!ラジオボタン選択時に入力
    def input_kensaku_entry(self):
        var=self.extractvar.get()
        self.kensaku_et.delete( 0, tk.END )
        self.kensaku_et.insert(0,globals()[f'syokuhinmei{var}'].get())
    def input_kensaku_entry_noinput(self,input_entry):
        var=self.extractvar.get()
        sss='!canvas2.!frame'
        if sss in str(input_entry):
            self.kensaku_et.delete( 0, tk.END )
            self.kensaku_et.insert(0,globals()[f'syokuhinmei{var}'].get()) 
    #!category登録有無確認
    def update_unregistered(self,event=None):
        for i in range(len(self.rfdf)+self.increase):
            self.unregistered(i)
        messagebox.showinfo("メッセージ", "categoryの登録有無を確認しました")
    #!今の行を削除
    def current_delete(self,event=None):
        var=self.extractvar.get()
        self.delete_row(var)
    #!削除（無効化）
    def delete_row(self,var):
        # var=self.extractvar.get()
        globals()[f'puru{var}'].config(state='disabled')
        globals()[f'daibunrui{var}'].config(state='disabled',foreground='black')
        globals()[f'tyubunrui{var}'].config(state='disabled',foreground='black')
        globals()[f'syokuhinmei{var}'].config(state='disabled',foreground='black')
        globals()[f'no{var}'].config(state='disabled')
        globals()[f'kazu{var}'].config(state='disabled')
        globals()[f'tani{var}'].config(state='disabled')
        globals()[f'kingaku{var}'].config(state='disabled')
        globals()[f'memo{var}'].config(state='disabled')
        globals()[f'unregistered{var}']["text"]=""
        self.active_list[var]=False
    
    #! マウスホールできるようにする
    def mouse_y_scroll(self, event):
        if event.delta > 0:
            self.raight_canvas.yview_scroll(-1, 'units')
        elif event.delta < 0:
            self.raight_canvas.yview_scroll(1, 'units')
    #* -------------------------------------------------------------------------------
    #* 描画
    #* -------------------------------------------------------------------------------
    def set_image(self):
        ''' 画像ファイルを開く '''
        if not self.filename:
            return
        # PIL.Imageで開く
        self.pil_image = Image.open(self.filename)
        # 画像全体に表示するようにアフィン変換行列を設定
        self.zoom_fit(self.pil_image.width, self.pil_image.height)
        # 画像の表示
        self.draw_image(self.pil_image)

    def draw_image(self, pil_image):     
        if pil_image == None:
            return   
        self.canvas.delete("all")
        # キャンバスのサイズ
        canvas_width = self.canvas.winfo_width()
        canvas_height = self.canvas.winfo_height()
        # キャンバスから画像データへのアフィン変換行列を求める
        #（表示用アフィン変換行列の逆行列を求める）
        mat_inv = np.linalg.inv(self.mat_affine)
        # PILの画像データをアフィン変換する
        dst = pil_image.transform(
                    (canvas_width, canvas_height),  # 出力サイズ
                    Image.AFFINE,                   # アフィン変換
                    tuple(mat_inv.flatten()),       # アフィン変換行列（出力→入力への変換行列）を一次元のタプルへ変換
                    Image.NEAREST,                  # 補間方法、ニアレストネイバー 
                    )
        # 表示用画像を保持
        self.image = ImageTk.PhotoImage(image=dst)
        # 画像の描画
        self.canvas.create_image(
                0, 0,               # 画像表示位置(左上の座標)
                anchor='nw',        # アンカー、左上が原点
                image=self.image    # 表示画像データ
                )
    def redraw_image(self):
        ''' 画像の再描画 '''
        if self.pil_image == None:
            return
        self.draw_image(self.pil_image)

    #!ファイル取得
    def file_path_get(self,event=None):
        self.statusbar["text"] = " fileGet"
        self.filename = filedialog.askopenfilename(
            filetypes=[("Image file", ".bmp .png .jpg .tif"), ("Bitmap", ".bmp"), ("PNG", ".png"), ("JPEG", ".jpg"), ("Tiff", ".tif") ],
            initialdir=r"C:\Users\f1351\Desktop\レシート"
            )
        #displayにファイル名を表示させる
        if self.filename!='':
            self.file_name_dis.delete( 0, tk.END )
        self.file_name_dis.insert(tk.END,self.filename)
        filename_len=int(len(self.filename)*1.25)
        if int(self.file_name_dis.cget('width'))<filename_len:
            self.file_name_dis.config(width=filename_len)
        self.set_image()
    
    #* -------------------------------------------------------------------------------
    #* マウスイベント
    #* -------------------------------------------------------------------------------
    def mouse_image(self):
        # self.master.bind("<Motion>", self.mouse_move)                       # MouseMove
        self.canvas.bind("<B1-Motion>", self.mouse_move_left)               # MouseMove（左ボタンを押しながら移動）
        self.canvas.bind("<Button-1>", self.mouse_down_left)                # MouseDown（左ボタン）
        self.canvas.bind("<Double-Button-1>", self.mouse_double_click_left) # MouseDoubleClick（左ボタン）
        self.canvas.bind("<MouseWheel>", self.mouse_wheel)

    def mouse_move_left(self, event):
        ''' マウスの左ボタンをドラッグ '''
        if self.pil_image == None:
            return
        self.translate(event.x - self.__old_event.x, event.y - self.__old_event.y)
        self.redraw_image() # 再描画
        self.__old_event = event

    def mouse_down_left(self, event):
        ''' マウスの左ボタンを押した '''
        self.__old_event = event

    def mouse_double_click_left(self, event):
        ''' マウスの左ボタンをダブルクリック '''
        if self.pil_image == None:
            return
        self.zoom_fit(self.pil_image.width, self.pil_image.height)
        self.redraw_image() # 再描画

    def mouse_wheel(self, event):
        ''' マウスホイールを回した '''
        if self.pil_image == None:
            return
        if (event.delta < 0):
            # 上に回転の場合、縮小
            self.scale_at(0.8, event.x, event.y)
        else:
            # 下に回転の場合、拡大
            self.scale_at(1.25, event.x, event.y)
        self.redraw_image() # 再描画

    #* -------------------------------------------------------------------------------
    #* 画像表示用アフィン変換
    #* -------------------------------------------------------------------------------

    def reset_transform(self):
        '''アフィン変換を初期化（スケール１、移動なし）に戻す'''
        self.mat_affine = np.eye(3) # 3x3の単位行列

    def translate(self, offset_x, offset_y):
        ''' 平行移動 '''
        mat = np.eye(3) # 3x3の単位行列
        mat[0, 2] = float(offset_x)
        mat[1, 2] = float(offset_y)
        self.mat_affine = np.dot(mat, self.mat_affine)

    def scale(self, scale:float):
        ''' 拡大縮小 '''
        mat = np.eye(3) # 単位行列
        mat[0, 0] = scale
        mat[1, 1] = scale

        self.mat_affine = np.dot(mat, self.mat_affine)

    def scale_at(self, scale:float, cx:float, cy:float):
        ''' 座標(cx, cy)を中心に拡大縮小 '''
        # 原点へ移動
        self.translate(-cx, -cy)
        # 拡大縮小
        self.scale(scale)
        # 元に戻す
        self.translate(cx, cy)

    def zoom_fit(self, image_width, image_height):
        '''画像をウィジェット全体に表示させる'''
        # キャンバスのサイズ
        canvas_width = self.canvas.winfo_width()
        canvas_height = self.canvas.winfo_height()
        if (image_width * image_height <= 0) or (canvas_width * canvas_height <= 0):
            return
        # アフィン変換の初期化
        self.reset_transform()
        scale = 1.0
        offsetx = 0.0
        offsety = 0.0

        if (canvas_width * image_height) > (image_width * canvas_height):
            # ウィジェットが横長（画像を縦に合わせる）
            scale = canvas_height / image_height
            # あまり部分の半分を中央に寄せる
            offsetx = (canvas_width - image_width * scale) / 2
        else:
            # ウィジェットが縦長（画像を横に合わせる）
            scale = canvas_width / image_width
            # あまり部分の半分を中央に寄せる
            offsety = (canvas_height - image_height * scale) / 2

        # 拡大縮小
        self.scale(scale)
        # あまり部分を中央に寄せる
        self.translate(offsetx, offsety)

if __name__ == '__main__':
    root=tk.Tk()
    app=Application(master=root)
    app.mainloop()